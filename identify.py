#import cognitive_face as CF
from azure.cognitiveservices.vision.face import FaceClient
from msrest.authentication import CognitiveServicesCredentials
from azure.cognitiveservices.vision.face.models import TrainingStatusType, Person, SnapshotObjectType, OperationStatusType
#import global_variables as global_var
import os, urllib
import sqlite3
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.cell import Cell
import time
import requests
from requests.packages.urllib3.exceptions import InsecureRequestWarning

requests.packages.urllib3.disable_warnings(InsecureRequestWarning)



#get current date
currentDate = time.strftime("%d_%m_%y")
wb = load_workbook(filename = "reports.xlsx")
sheet = wb['Cse16']

def getDateColumn():
	for i in range(1, len(list(sheet.rows)[0]) + 1):
		col = get_column_letter(i)
		if sheet['%s%s'% (col,'1')].value == currentDate:
			return col

Key ='33511ab23d274155b6a06966bf896b19'


ENDPOINT = 'https://recogination.cognitiveservices.azure.com/'
face_client = FaceClient(ENDPOINT,CognitiveServicesCredentials(Key))
person_group_id='face1'

connect = sqlite3.connect("Face-DataBase")


attend = [0 for i in range(60)]	

currentDir = os.path.dirname(os.path.abspath(__file__))
directory = os.path.join(currentDir, 'Cropped_faces')
for filename in os.listdir(directory):
	if filename.endswith(".jpg"):
		print(filename)
		img_data = open(os.path.join(directory,filename), "rb")
		res = face_client.face.detect_with_stream(img_data)
		#print("Res = {}".format(res))

		if len(res) < 1:
			print("No face detected.")
			continue
			
		faceIds = []
		for face in res:
			faceIds.append(face.face_id)
		res = face_client.face.identify(faceIds, person_group_id)
		#print(filename)
		#print("res = {}".format(res))

		for face  in res:
			if not face.candidates:
				print("Unknown")
			else:
				personId = face.candidates[0].person_id
				#print("personid = {}".format(personId))
				
				cur = connect.execute("SELECT * FROM Students WHERE personID = (?)", (personId,))
				#print("cur = {}".format(cur))
				for row in cur:
					#print("aya")
					print("row = {}".format(row))
					attend[int(row[0])] += 1
					print("---------- " + row[1] + " recognized ----------")
		time.sleep(6)
		
for row in range(2, len(list(sheet.columns)[0]) + 1):
	rn = sheet.cell(row = row, column  =1).value
	if rn is not None:
		#print("rn = {}".format(rn))
		rn = rn[-2:]
		if attend[int(rn)] != 0:
			col = getDateColumn()
			#print("col = {}".format(col))
			sheet['%s%s' % (col, str(row))] ="P"
		

wb.save(filename = "reports.xlsx")	 	
#currentDir = os.path.dirname(os.path.abspath(__file__))
#imgurl = urllib.pathname2url(os.path.join(currentDir, "1.jpg"))
#res = CF.face.detect(imgurl)
#faceIds = []
#for face in res:
 #   faceIds.append(face['faceId'])

#res = CF.face.identify(faceIds,personGroupId)
# for face in res:
#     personName = CF.person.get(personGroupId, face['candidates']['personId'])
#     print personName
#print res